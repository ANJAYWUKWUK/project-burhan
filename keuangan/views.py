from django.db import models 
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Bank, Transaksi, Hutang, Piutang, Kategori, HutangPiutang,PembayaranSPP, Siswa, LaporanKeuangan, TabunganSiswa , PenarikanTabungan,DSP,DSPCicilan, PPDB, PPDBCicilan
from .forms import TransaksiForm, HutangForm, PiutangForm, BankForm, KategoriForm, HutangPiutangForm,BuktiPembayaranForm,PilihKelasForm, PilihBulanForm,DSPForm,DSPCicilanForm,PPDBForm, PPDBCicilanForm
from django.contrib.auth import authenticate, login
from django.views.decorators.http import require_POST
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import logout
from django.contrib import messages
from .forms import KategoriForm, BankForm
from django.db.models import Sum
from django.utils import timezone
from openpyxl.styles import Font
from django.db.models.functions import TruncDay, TruncMonth, TruncYear
import json
import calendar
from datetime import timedelta
from django.conf import settings
from django.contrib.auth import authenticate, login
from django.core.mail import send_mail
from django.http import HttpResponse
import openpyxl
from io import BytesIO
from .functions import generate_tagihan_spp
from datetime import datetime
from django.core.files.storage import FileSystemStorage
from django.utils.datastructures import MultiValueDictKeyError
from django.http import JsonResponse
from django.db import transaction   
from openpyxl.utils import get_column_letter

def login_view(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)  # HARUS ADA INI
            return redirect('dashboard_redirect')  # pastikan 'dashboard' ini ada di urls.py
    else:
        form = AuthenticationForm()
    return render(request, 'keuangan/login.html', {'form': form})

@login_required
def dashboard_redirect(request):
    user = request.user
    if user.groups.filter(name='Admin').exists():
        return redirect('dashboard_admin')
    elif user.groups.filter(name='Siswa').exists():
        return redirect('dashboard_siswa')
    return redirect('login')


@login_required
def dashboard(request):
    user = request.user
    if not user.groups.filter(name='Admin').exists():
        return redirect('dashboard_siswa') # Asumsikan Anda punya URL ini untuk siswa
    
    today = timezone.now().date()
    this_month_start = today.replace(day=1)
    this_year_start = today.replace(month=1, day=1)
    
    # Ambil data dari model LaporanKeuangan
    # Perhatikan: ini akan membaca agregasi yang sudah disimpan oleh signals
    
    # Pemasukan Hari Ini
    laporan_hari_ini = LaporanKeuangan.objects.filter(tanggal=today).first()
    pemasukan_hari_ini = laporan_hari_ini.pemasukan if laporan_hari_ini else 0
    pengeluaran_hari_ini = laporan_hari_ini.pengeluaran if laporan_hari_ini else 0

    # Pemasukan Bulan Ini
    pemasukan_bulan_ini = LaporanKeuangan.objects.filter(
        tanggal__gte=this_month_start, tanggal__lte=today
    ).aggregate(Sum('pemasukan'))['pemasukan__sum'] or 0

    # Pemasukan Tahun Ini
    pemasukan_tahun_ini = LaporanKeuangan.objects.filter(
        tanggal__gte=this_year_start, tanggal__lte=today
    ).aggregate(Sum('pemasukan'))['pemasukan__sum'] or 0
    
    # Total Pemasukan Keseluruhan
    total_pemasukan = LaporanKeuangan.objects.aggregate(
        Sum('pemasukan')
    )['pemasukan__sum'] or 0

    # Pengeluaran Bulan Ini
    pengeluaran_bulan_ini = LaporanKeuangan.objects.filter(
        tanggal__gte=this_month_start, tanggal__lte=today
    ).aggregate(Sum('pengeluaran'))['pengeluaran__sum'] or 0

    # Pengeluaran Tahun Ini
    pengeluaran_tahun_ini = LaporanKeuangan.objects.filter(
        tanggal__gte=this_year_start, tanggal__lte=today
    ).aggregate(Sum('pengeluaran'))['pengeluaran__sum'] or 0
    
    # Total Pengeluaran Keseluruhan
    total_pengeluaran = LaporanKeuangan.objects.aggregate(
        Sum('pengeluaran')
    )['pengeluaran__sum'] or 0

    # Data Bulanan untuk Grafik (dari model LaporanKeuangan)
    pemasukan_bulanan_raw = LaporanKeuangan.objects.filter(
        tanggal__year=this_year_start.year
    ).values('tanggal__month').annotate(total=Sum('pemasukan')).order_by('tanggal__month')

    pengeluaran_bulanan_raw = LaporanKeuangan.objects.filter(
        tanggal__year=this_year_start.year
    ).values('tanggal__month').annotate(total=Sum('pengeluaran')).order_by('tanggal__month')

    pemasukan_chart_data = [0] * 12
    pengeluaran_chart_data = [0] * 12

    for entry in pemasukan_bulanan_raw:
        month_index = entry['tanggal__month'] - 1
        pemasukan_chart_data[month_index] = float(entry['total'])

    for entry in pengeluaran_bulanan_raw:
        month_index = entry['tanggal__month'] - 1
        pengeluaran_chart_data[month_index] = float(entry['total'])

    # --- Metrik Tambahan (optional, jika Anda ingin menampilkan di dashboard) ---
    # SPP
    spp_total_terbayar = PembayaranSPP.objects.filter(
        status_bayar='lunas'
    ).aggregate(Sum('jumlah_bayar'))['jumlah_bayar__sum'] or 0
    spp_belum_lunas_count = PembayaranSPP.objects.filter(
        status_bayar='belum lunas'
    ).count()

    # DSP
    total_dsp_tagihan = DSP.objects.aggregate(Sum('total_tagihan'))['total_tagihan__sum'] or 0
    total_dsp_terbayar_semua = DSP.objects.annotate(
        terbayar=Sum('dspcicilan__jumlah')
    ).aggregate(total=Sum('terbayar'))['total'] or 0
    total_dsp_sisa = total_dsp_tagihan - total_dsp_terbayar_semua
    dsp_belum_lunas_count = DSP.objects.filter(status='belum lunas').count()

    # PPDB
    total_ppdb_tagihan = PPDB.objects.aggregate(Sum('total_tagihan'))['total_tagihan__sum'] or 0
    total_ppdb_terbayar_semua = PPDB.objects.annotate(
        terbayar=Sum('ppdbcicilan__jumlah')
    ).aggregate(total=Sum('terbayar'))['total'] or 0
    total_ppdb_sisa = total_ppdb_tagihan - total_ppdb_terbayar_semua
    ppdb_belum_lunas_count = PPDB.objects.filter(status='Belum Lunas').count()
    
    # Tabungan Siswa
    total_tabungan_masuk = TabunganSiswa.objects.filter(tarik=False).aggregate(Sum('nominal'))['nominal__sum'] or 0
    total_penarikan_tabungan = PenarikanTabungan.objects.aggregate(Sum('jumlah_ditarik'))['jumlah_ditarik__sum'] or 0
    saldo_tabungan_bersih = total_tabungan_masuk - total_penarikan_tabungan

    # Total Siswa
    total_siswa = Siswa.objects.count()


    context = {
        'pemasukan_hari_ini': pemasukan_hari_ini,
        'pemasukan_bulan_ini': pemasukan_bulan_ini,
        'pemasukan_tahun_ini': pemasukan_tahun_ini,
        'total_pemasukan': total_pemasukan,

        'pengeluaran_hari_ini': pengeluaran_hari_ini,
        'pengeluaran_bulan_ini': pengeluaran_bulan_ini,
        'pengeluaran_tahun_ini': pengeluaran_tahun_ini,
        'total_pengeluaran': total_pengeluaran,

        'pemasukan_bulanan': json.dumps(pemasukan_chart_data),
        'pengeluaran_bulanan': json.dumps(pengeluaran_chart_data),

        'spp_total_terbayar': spp_total_terbayar,
        'spp_belum_lunas_count': spp_belum_lunas_count,
        # 'spp_jatuh_tempo_mendekat': spp_jatuh_tempo_mendekat, # Butuh logic lebih lanjut untuk ini

        'total_dsp_tagihan': total_dsp_tagihan,
        'total_dsp_terbayar_semua': total_dsp_terbayar_semua,
        'total_dsp_sisa': total_dsp_sisa,
        'dsp_belum_lunas_count': dsp_belum_lunas_count,

        'total_ppdb_tagihan': total_ppdb_tagihan,
        'total_ppdb_terbayar_semua': total_ppdb_terbayar_semua,
        'total_ppdb_sisa': total_ppdb_sisa,
        'ppdb_belum_lunas_count': ppdb_belum_lunas_count,

        'total_tabungan_masuk': total_tabungan_masuk,
        'total_penarikan_tabungan': total_penarikan_tabungan,
        'saldo_tabungan_bersih': saldo_tabungan_bersih,

        'total_siswa': total_siswa,
    }
    return render(request, 'keuangan/dashboard.html', context)

@login_required
def dashboard_siswa(request):
    user = request.user
    # Cek apakah user termasuk grup Siswa
    if not user.groups.filter(name='Siswa').exists():
        return redirect('dashboard_admin')  # atau redirect ke mana pun kalau bukan siswa

    tagihan_saya = PembayaranSPP.objects.filter(siswa__user=request.user)
    return render(request, 'keuangan/siswa/dashboard_siswa.html', {'tagihan': tagihan_saya})

def logout_view(request):
    logout(request)
    return redirect('login')  # Ganti 'login' dengan halaman tujuan setelah logout[



@login_required
def kelola_pembayaran(request):
    kelas_list = Siswa.objects.values_list('kelas', flat=True).distinct()

    kelas_terpilih = request.GET.get('kelas')
    siswa_id = request.GET.get('siswa_id')
    search_nama = request.GET.get('search_nama')

    pembayaran_list = None
    siswa_list = []
    siswa_terpilih = None

    if kelas_terpilih:
        siswa_list = Siswa.objects.filter(kelas=kelas_terpilih)

        if siswa_id:
            siswa_terpilih = get_object_or_404(Siswa, id=siswa_id)
            pembayaran_list = PembayaranSPP.objects.filter(siswa=siswa_terpilih).order_by('bulan')

    if request.method == 'POST':
        pembayaran_id = request.POST.get('pembayaran_id')
        pembayaran = get_object_or_404(PembayaranSPP, id=pembayaran_id)

        pembayaran.status_bayar = 'lunas'
        pembayaran.save()

        # redirect tetap jaga query string supaya tetap di halaman yang sama
        redirect_url = request.path
        query_params = []
        if kelas_terpilih:
            query_params.append(f'kelas={kelas_terpilih}')
        if siswa_id:
            query_params.append(f'siswa_id={siswa_id}')
        if search_nama:
            query_params.append(f'search_nama={search_nama}')
        if query_params:
            redirect_url += '?' + '&'.join(query_params)

        return redirect(redirect_url)

    return render(request, 'keuangan/admin/kelola_pembayaran.html', {
        'kelas_list': kelas_list,
        'kelas_terpilih': kelas_terpilih,
        'siswa_list': siswa_list,
        'siswa_terpilih': siswa_terpilih,
        'pembayaran_list': pembayaran_list,
        'search_nama': search_nama,
    })



def is_siswa(user):
    return hasattr(user, 'siswa')
# Siswa View: Lihat tagihan pribadi & upload bukti
@user_passes_test(is_siswa, login_url='/login/')
@login_required
def tagihan_spp(request):
    # Ambil data siswa berdasarkan user yang login
    siswa = get_object_or_404(Siswa, user=request.user)

    # Ambil semua tagihan siswa tersebut
    pembayaran_list = PembayaranSPP.objects.filter(siswa=siswa).order_by('bulan')

    if request.method == 'POST':
        pembayaran_id = request.POST.get('pembayaran_id')
        pembayaran = get_object_or_404(PembayaranSPP, id=pembayaran_id, siswa=siswa)

        form = BuktiPembayaranForm(request.POST, request.FILES, instance=pembayaran)
        if form.is_valid():
            # Simpan bukti pembayaran dan ubah status jadi "pending" atau lainnya
            pembayaran.status = 'pending'  # optional: kalau lo ada field status
            form.save()

             # Update laporan keuangan
            laporan = LaporanKeuangan.objects.latest('tanggal')
            laporan.update_pemasukan(pembayaran.jumlah_bayar)

            send_mail(
                subject=f'Bukti Pembayaran SPP Masuk - {pembayaran.siswa.nama}',
                message=f'Siswa {pembayaran.siswa.nama} telah mengupload bukti pembayaran untuk bulan {pembayaran.bulan}. Jumlah yang dibayar: {pembayaran.jumlah_bayar}.',
                from_email=settings.EMAIL_HOST_USER,
                recipient_list=[settings.ADMIN_EMAIL],
                fail_silently=False,
            )

            return redirect('tagihan_spp')
    else:
        form = BuktiPembayaranForm()

    return render(request, 'keuangan/siswa/tagihan_spp.html', {
        'pembayaran_list': pembayaran_list,
        'form': form
    })

def test_email(request):
    send_mail(
        subject='[TEST] Bukti Pembayaran Masuk',
        message='Ini adalah simulasi notifikasi: Siswa Fulan sudah upload bukti pembayaran.',
        from_email=settings.EMAIL_HOST_USER,
        recipient_list=[settings.ADMIN_EMAIL],
        fail_silently=False,
    )
    return HttpResponse("Email test berhasil dikirim.")

@login_required
def export_pembayaran_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pembayaran SPP"

    # Header
    ws.append(['Nama Siswa', 'Kelas', 'Bulan', 'Jumlah Bayar', 'Status', 'Tanggal Bayar'])

    # Data
    pembayaran_list = PembayaranSPP.objects.all().order_by('siswa__nama')
    for bayar in pembayaran_list:
        ws.append([
            bayar.siswa.nama,
            bayar.siswa.kelas,
            bayar.get_bulan_display(),
            bayar.jumlah_bayar,
            bayar.status_bayar,
            bayar.tanggal_bayar.strftime("%d-%m-%Y") if bayar.tanggal_bayar else '-'
        ])

    # Simpan workbook ke dalam memory (BytesIO)
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Response download file
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=pembayaran_spp.xlsx'
    return response

def daftar_siswa(request):
    siswa_list = Siswa.objects.all()

    if request.method == "POST":
        siswa_id = request.POST.get("siswa_id")
        kelas_baru = request.POST.get("kelas_baru")
        siswa = get_object_or_404(Siswa, id=siswa_id)
        siswa.kelas = kelas_baru
        siswa.save()
        return redirect('daftar_siswa')  # ganti dengan nama URL kamu

    return render(request, 'keuangan/admin/daftar_siswa.html', {'siswa_list': siswa_list})

def generate_tagihan(request, siswa_id):
    siswa = Siswa.objects.get(id=siswa_id)
    generate_tagihan_spp(siswa)
    return redirect('siswa_detail', siswa_id=siswa.id)

@login_required
def form_dsp(request, dsp_id=None):
    if dsp_id:
        dsp = get_object_or_404(DSP, id=dsp_id)
        judul = 'Edit DSP'
    else:
        dsp = None
        judul = 'Tambah DSP'

    if request.method == 'POST':
        form = DSPForm(request.POST, instance=dsp)
        if form.is_valid():
            form.save()
            messages.success(request, f'Data DSP berhasil {"diperbarui" if dsp_id else "ditambahkan"}.')
            return redirect('dsp_list')
    else:
        form = DSPForm(instance=dsp)

    return render(request, 'keuangan/admin/form_dsp.html', {'form': form, 'judul': judul})


@login_required
def form_cicilan_dsp(request, dsp_id, cicilan_id=None):
    dsp = get_object_or_404(DSP, id=dsp_id)
    
    if cicilan_id:
        cicilan = get_object_or_404(DSPCicilan, id=cicilan_id, dsp=dsp)
        judul = f"Edit Cicilan DSP - {dsp.siswa.nama}"
    else:
        cicilan = None
        judul = f"Tambah Cicilan DSP - {dsp.siswa.nama}"

    if request.method == 'POST':
        form = DSPCicilanForm(request.POST, request.FILES, instance=cicilan)
        if form.is_valid():
            new_cicilan = form.save(commit=False)
            new_cicilan.dsp = dsp
            new_cicilan.save()

            # ✅ Update total terbayar & status otomatis
            total_terbayar = DSPCicilan.objects.filter(dsp=dsp).aggregate(Sum('jumlah'))['jumlah__sum'] or 0
            dsp.total_terbayar = total_terbayar
            dsp.sisa_tagihan = dsp.total_tagihan - total_terbayar

            # ✅ Tambahkan logika manual override status jika admin ingin
            if 'set_status_lunas' in request.POST:
                dsp.status = "Lunas"
            elif 'set_status_belum' in request.POST:
                dsp.status = "Belum Lunas"
            else:
                dsp.status = "Lunas" if dsp.sisa_tagihan <= 0 else "Belum Lunas"

            dsp.save()

            messages.success(request, f'Cicilan berhasil {"diperbarui" if cicilan_id else "ditambahkan"}.')
            return redirect('dsp_cicilan_list', dsp_id=dsp.id)
    else:
        form = DSPCicilanForm(instance=cicilan)

    return render(request, 'keuangan/admin/dsp_cicilan_list.html', {
        'form': form,
        'judul': judul,
        'dsp': dsp,
        'cicilan': cicilan,
    })

# views.py

def dsp_cicilan_list(request, dsp_id):
    dsp = get_object_or_404(DSP, id=dsp_id)
    cicilan_list = DSPCicilan.objects.filter(dsp=dsp).order_by('tanggal')

    return render(request, 'keuangan/admin/cicilan_list.html', {
        'dsp': dsp,
        'cicilan_list': cicilan_list,
    })


from django.utils import timezone

@login_required
def dsp_siswa_detail(request):
    siswa = request.user.siswa  # Pastikan user terkait dengan model Siswa
    dsp = get_object_or_404(DSP, siswa=siswa)
    cicilan_list = DSPCicilan.objects.filter(dsp=dsp).order_by('tanggal')

    pesan = ""
    if request.method == "POST":
        jumlah = request.POST.get("jumlah")
        bukti = request.FILES.get("bukti")

        if jumlah:
            DSPCicilan.objects.create(
                dsp=dsp,
                tanggal=timezone.now().date(),
                jumlah=jumlah,
                bukti_pembayaran=bukti
            )

            # ✅ Gunakan aggregate(Sum(...)) dengan benar
            total_terbayar = DSPCicilan.objects.filter(dsp=dsp).aggregate(Sum('jumlah'))['jumlah__sum'] or 0

            dsp.total_terbayar = total_terbayar
            dsp.sisa_tagihan = dsp.total_tagihan - total_terbayar
            dsp.status = "Lunas" if dsp.sisa_tagihan <= 0 else "Belum Lunas"
            dsp.save()

            pesan = "Cicilan berhasil dikirim."

    return render(request, 'keuangan/siswa/dsp_siswa_detail.html', {
        'dsp': dsp,
        'cicilan_list': cicilan_list,
        'pesan': pesan
    })


@login_required
def dsp_list(request):
    dsps = DSP.objects.select_related('siswa')
    return render(request, 'keuangan/admin/daftar_dsp.html', {'dsps': dsps})

def dsp_delete(request, dsp_id):
    dsp = get_object_or_404(DSP, id=dsp_id)
    dsp.delete()
    messages.success(request, f'Data DSP untuk {dsp.siswa.nama} berhasil dihapus.')
    return redirect('dsp_list')

@login_required
def ubah_status_dsp(request, dsp_id, status):
    dsp = get_object_or_404(DSP, id=dsp_id)

    if status not in ['belum lunas', 'lunas']:
        messages.error(request, 'Status tidak valid.')
        return redirect('dsp_cicilan_list', dsp_id=dsp_id)

    dsp.status = status
    dsp.save()
    messages.success(request, f'Status DSP berhasil diubah menjadi {status.upper()}')
    return redirect('dsp_cicilan_list', dsp_id=dsp_id)

@login_required
def ppdb_list(request):
    ppdbs = PPDB.objects.select_related('siswa')
    return render(request, 'keuangan/admin/ppdb_list.html', {'ppdbs': ppdbs})

@login_required
def form_ppdb(request, ppdb_id=None):
    if ppdb_id:
        ppdb = get_object_or_404(PPDB, id=ppdb_id)
        judul = 'Edit PPDB'
    else:
        ppdb = None
        judul = 'Tambah PPDB'

    if request.method == 'POST':
        form = PPDBForm(request.POST, instance=ppdb)
        if form.is_valid():
            ppdb_obj = form.save(commit=False)
            if not ppdb_id:
                ppdb_obj.total_terbayar = 0
                ppdb_obj.sisa_tagihan = ppdb_obj.total_tagihan
            ppdb_obj.save()
            messages.success(request, f'Data PPDB berhasil {"diperbarui" if ppdb_id else "ditambahkan"}.')
            return redirect('ppdb_list')
    else:
        form = PPDBForm(instance=ppdb)

    return render(request, 'keuangan/admin/form_ppdb.html', {'form': form, 'judul': judul})

# Fungsi helper untuk update total pembayaran PPDB
def update_ppdb_total(ppdb):
    total = PPDBCicilan.objects.filter(ppdb=ppdb).aggregate(Sum('jumlah'))['jumlah__sum'] or 0
    ppdb.total_terbayar = total
    ppdb.sisa_tagihan = ppdb.total_tagihan - total
    ppdb.status = "Lunas" if ppdb.sisa_tagihan <= 0 else "Belum Lunas"
    ppdb.save()

# View detail PPDB
@login_required
def ppdb_detail(request, ppdb_id):
    ppdb = get_object_or_404(PPDB, id=ppdb_id)
    ppdb.refresh_from_db()  # Penting untuk ambil data paling update dari database
    cicilan_list = PPDBCicilan.objects.filter(ppdb=ppdb).order_by('tanggal')

    return render(request, 'keuangan/admin/ppdb_detail.html', {
        'ppdb': ppdb,
        'cicilan_list': cicilan_list
    })

# View tambah/edit cicilan PPDB
@login_required
def form_cicilan_ppdb(request, ppdb_id, cicilan_id=None):
    ppdb = get_object_or_404(PPDB, id=ppdb_id)
    cicilan = get_object_or_404(PPDBCicilan, id=cicilan_id, ppdb=ppdb) if cicilan_id else None
    judul = f"{'Edit' if cicilan else 'Tambah'} Cicilan PPDB - {ppdb.siswa.nama}"

    if request.method == 'POST':
        form = PPDBCicilanForm(request.POST, request.FILES, instance=cicilan)
        if form.is_valid():
            new_cicilan = form.save(commit=False)
            new_cicilan.ppdb = ppdb
            new_cicilan.save()

            # Update total PPDB
            update_ppdb_total(ppdb)

            messages.success(request, 'Cicilan berhasil disimpan.')
            return redirect('ppdb_detail', ppdb_id=ppdb.id)
    else:
        form = PPDBCicilanForm(instance=cicilan)

    return render(request, 'keuangan/admin/form_cicilan_ppdb.html', {
        'form': form,
        'judul': judul,
        'ppdb': ppdb
    })

@login_required
def ppdb_delete(request, ppdb_id):
    ppdb = get_object_or_404(PPDB, id=ppdb_id)
    ppdb.delete()
    messages.success(request, f'Data PPDB untuk {ppdb.siswa.nama} berhasil dihapus.')
    return redirect('ppdb_list')

@login_required
def siswa_ppdb_detail(request):
    siswa = request.user.siswa
    ppdb = get_object_or_404(PPDB, siswa=siswa)
    cicilan_list = PPDBCicilan.objects.filter(ppdb=ppdb).order_by('-tanggal')

    return render(request, 'keuangan/siswa/ppdb_siswa_detail.html', {
        'ppdb': ppdb,
        'cicilan_list': cicilan_list,
    })

@login_required
def siswa_ppdb_cicilan_tambah(request, ppdb_id):
    siswa = request.user.siswa
    ppdb = get_object_or_404(PPDB, id=ppdb_id, siswa=siswa)

    if request.method == 'POST':
        form = PPDBCicilanForm(request.POST, request.FILES)
        if form.is_valid():
            cicilan = form.save(commit=False)

            # Jumlah cicilan baru
            jumlah_baru = cicilan.jumlah

            # Hitung total terbayar sejauh ini
            total_terbayar_sekarang = PPDBCicilan.objects.filter(ppdb=ppdb).aggregate(Sum('jumlah'))['jumlah__sum'] or 0

            sisa = ppdb.total_tagihan - total_terbayar_sekarang

            if jumlah_baru > sisa:
                form.add_error('jumlah', f"Jumlah cicilan melebihi sisa tagihan (sisa: Rp{sisa:,})")
            else:
                cicilan.ppdb = ppdb
                cicilan.tanggal = timezone.now().date()
                cicilan.save()

                # Hitung ulang total terbayar & status
                total_terbayar = PPDBCicilan.objects.filter(ppdb=ppdb).aggregate(Sum('jumlah'))['jumlah__sum'] or 0
                ppdb.total_terbayar = total_terbayar
                ppdb.sisa_tagihan = max(ppdb.total_tagihan - total_terbayar, 0)
                ppdb.status = "Lunas" if ppdb.sisa_tagihan <= 0 else "Belum Lunas"
                ppdb.save()

                return redirect('siswa_ppdb_detail')

    else:
        form = PPDBCicilanForm()

    return render(request, 'keuangan/siswa/form_cicilan_siswa.html', {
        'form': form,
        'ppdb': ppdb
    })

def get_siswa_by_kelas(request):
    kelas = request.GET.get("kelas")
    siswa_list = Siswa.objects.filter(kelas=kelas).values("id", "nama")
    return JsonResponse(list(siswa_list), safe=False)

def get_pembayaran_by_siswa(request):
    siswa_id = request.GET.get("siswa_id")
    pembayaran = PembayaranSPP.objects.filter(siswa_id=siswa_id).values(
        "id", "bulan", "jumlah_bayar", "status_bayar", "tanggal_bayar", "bukti_pembayaran"
    )
    return JsonResponse(list(pembayaran), safe=False)

from django.shortcuts import render, redirect
from django.contrib import messages
from django.db.models import Sum
from django.db import transaction
from datetime import datetime
import calendar

from .models import Siswa, TabunganSiswa, PenarikanTabungan
from .forms import PilihKelasForm, PilihBulanForm


def tabungan_view(request):
    mode = request.GET.get('mode', 'input')
    context = {'mode': mode}

    if mode == 'input':
        kelas_form = PilihKelasForm(request.POST or None)
        bulan_form = PilihBulanForm(request.POST or None)
        siswa_list, tanggal_list, checked_data = [], [], {}
        selected_kelas = request.POST.get('kelas') or request.GET.get('kelas')
        selected_bulan = request.POST.get('bulan') or request.GET.get('bulan')
        selected_tahun = request.POST.get('tahun') or request.GET.get('tahun')

        if selected_kelas and selected_bulan and selected_tahun:
            try:
                selected_bulan = int(selected_bulan)
                selected_tahun = int(selected_tahun)
                siswa_list = Siswa.objects.filter(kelas=selected_kelas)
                _, last_day = calendar.monthrange(selected_tahun, selected_bulan)
                tanggal_list = list(range(1, last_day + 1))
            except ValueError:
                messages.error(request, "Bulan dan Tahun harus berupa angka.")
                return redirect('kelola_tabungan')

            if request.method == 'POST' and 'simpan' in request.POST:
                for siswa in siswa_list:
                    checked_data[siswa.id] = request.POST.getlist(f'tanggal_{siswa.id}[]')

                for siswa in siswa_list:
                    tanggal_terpilih = {
                        datetime(selected_tahun, selected_bulan, int(t)).date()
                        for t in checked_data.get(siswa.id, [])
                    }

                    existing = TabunganSiswa.objects.filter(
                        siswa=siswa, tanggal__month=selected_bulan, tanggal__year=selected_tahun
                    )
                    existing_tanggal = set(e.tanggal for e in existing)

                    TabunganSiswa.objects.filter(
                        siswa=siswa, tanggal__in=existing_tanggal - tanggal_terpilih
                    ).delete()

                    for tgl in tanggal_terpilih - existing_tanggal:
                        TabunganSiswa.objects.get_or_create(
                            siswa=siswa, tanggal=tgl, defaults={'nominal': 5000}
                        )

                messages.success(request, "Tabungan berhasil disimpan.")
            else:
                tabungan = TabunganSiswa.objects.filter(
                    siswa__in=siswa_list, tanggal__month=selected_bulan, tanggal__year=selected_tahun
                )
                for item in tabungan:
                    key = f"{item.siswa.id}_{item.tanggal.month}_{item.tanggal.year}"
                    checked_data.setdefault(key, []).append(item.tanggal.strftime('%Y-%m-%d'))

        context.update({
            'kelas_form': kelas_form,
            'bulan_form': bulan_form,
            'siswa_list': siswa_list,
            'tanggal_list': tanggal_list,
            'selected_kelas': selected_kelas,
            'selected_bulan': selected_bulan,
            'selected_tahun': selected_tahun,
            'checked_data': checked_data,
        })

    elif mode == 'riwayat':
        selected_kelas = request.GET.get('kelas', '')
        selected_siswa = request.GET.get('siswa', '')
        start_date = request.GET.get('start_date', '')
        end_date = request.GET.get('end_date', '')
        siswa_list = Siswa.objects.filter(kelas=selected_kelas) if selected_kelas else Siswa.objects.all()
        tabungan = TabunganSiswa.objects.none()
        total_dict = {}

        if selected_kelas and start_date and end_date:
            tabungan = TabunganSiswa.objects.filter(
                siswa__kelas=selected_kelas,
                tanggal__range=[start_date, end_date]
            )
            if selected_siswa:
                tabungan = tabungan.filter(siswa__id=selected_siswa)

            total_tabungan = tabungan.values('siswa').annotate(total=Sum('nominal'))
            total_dict = {item['siswa']: item['total'] for item in total_tabungan}

        context.update({
            'tabungan_list': tabungan.order_by('tanggal'),
            'siswa_list': siswa_list,
            'kelas_list': Siswa.objects.values('kelas').distinct(),
            'start_date': start_date,
            'end_date': end_date,
            'selected_kelas': selected_kelas,
            'selected_siswa': selected_siswa,
            'total_tabungan': total_dict,
        })

    elif mode == 'penarikan':
        selected_kelas = request.GET.get('kelas', '')
        selected_siswa = request.GET.get('siswa', '')
        siswa_list = Siswa.objects.filter(kelas=selected_kelas) if selected_kelas else Siswa.objects.all()
        siswa_tabungan = TabunganSiswa.objects.select_related('siswa') \
            .filter(siswa__in=siswa_list, tarik=False) \
            .values('siswa__id', 'siswa__nama') \
            .annotate(total_tabungan=Sum('nominal'))

        if selected_siswa:
            siswa_tabungan = siswa_tabungan.filter(siswa__id=selected_siswa)

        if request.method == 'POST':
            tanggal_penarikan = request.POST.get('tanggal_penarikan')
            siswa_ditarik_ids = request.POST.getlist('siswa_ditarik')

            if not tanggal_penarikan:
                messages.error(request, "Tanggal penarikan wajib diisi.")
                return redirect('kelola_tabungan')

            success_count = 0
            with transaction.atomic():
                for siswa_id in siswa_ditarik_ids:
                    siswa = Siswa.objects.filter(id=siswa_id).first()
                    if not siswa:
                        continue

                    nominal_str = request.POST.get(f'nominal_{siswa_id}', '0').replace('.', '').replace(',', '')
                    try:
                        nominal = int(nominal_str)
                    except ValueError:
                        continue

                    total_tabungan = TabunganSiswa.objects.filter(
                        siswa=siswa, tarik=False
                    ).aggregate(total=Sum('nominal'))['total'] or 0

                    if nominal <= 0 or nominal > total_tabungan:
                        continue

                    PenarikanTabungan.objects.create(
                        siswa=siswa,
                        tanggal_penarikan=tanggal_penarikan,
                        jumlah_ditarik=nominal,
                        keterangan=f"Penarikan sebesar Rp{nominal:,}"
                    )

                    tabungan_entries = TabunganSiswa.objects.filter(
                        siswa=siswa, tarik=False
                    ).order_by('tanggal')
                    sisa_tarik = nominal
                    for entry in tabungan_entries:
                        if sisa_tarik <= 0:
                            break
                        if entry.nominal <= sisa_tarik:
                            sisa_tarik -= entry.nominal
                            entry.nominal = 0
                            entry.tarik = True
                        else:
                            entry.nominal -= sisa_tarik
                            sisa_tarik = 0
                        entry.save()
                    success_count += 1

            messages.success(request, f"{success_count} penarikan berhasil diproses.")
            return redirect('kelola_tabungan')

        context.update({
            'siswa_tabungan': siswa_tabungan,
            'kelas_list': Siswa.objects.values_list('kelas', flat=True).distinct(),
            'siswa_list': siswa_list,
            'selected_kelas': selected_kelas,
            'selected_siswa': selected_siswa,
        })

    return render(request, 'keuangan/admin/tabungan.html', context)

@login_required
def tabungan_siswa_view(request):
    siswa = getattr(request.user, 'siswa', None)
    if not siswa:
        return redirect('dashboard')  # atau sesuaikan dengan redirect lain

    # ❗ Ambil hanya yang belum ditarik
    tabungan_masuk = TabunganSiswa.objects.filter(siswa=siswa, tarik=False)
    penarikan = PenarikanTabungan.objects.filter(siswa=siswa)

    total_setoran = sum(t.nominal for t in tabungan_masuk)
    total_penarikan = sum(p.jumlah_ditarik for p in penarikan)

    # ❗ Biar tidak minus
    saldo = max(total_setoran - total_penarikan, 0)

    context = {
        'siswa': siswa,
        'tabungan_masuk': tabungan_masuk,
        'penarikan': penarikan,
        'saldo': saldo,
    }

    return render(request, 'keuangan/siswa/tabungan_siswa.html', context)

def laporan_keuangan_view(request):
    tanggal_hari_ini = timezone.now().date()
    kategori = request.GET.get('kategori')
    export = request.GET.get('export')

    # Ambil SEMUA transaksi
    spp = PembayaranSPP.objects.filter(status_bayar='lunas')
    dsp = DSPCicilan.objects.all()
    ppdb = PPDBCicilan.objects.all()
    tabungan = TabunganSiswa.objects.all()
    penarikan = PenarikanTabungan.objects.all()

    # Hitung total pemasukan & pengeluaran
    total_masuk = (
        (spp.aggregate(Sum('jumlah_bayar'))['jumlah_bayar__sum'] or 0) +
        (dsp.aggregate(Sum('jumlah'))['jumlah__sum'] or 0) +
        (ppdb.aggregate(Sum('jumlah'))['jumlah__sum'] or 0) +
        (tabungan.aggregate(Sum('nominal'))['nominal__sum'] or 0)
    )
    total_keluar = (penarikan.aggregate(Sum('jumlah_ditarik'))['jumlah_ditarik__sum'] or 0)

    # Export ke Excel berdasarkan kategori
    if export == 'excel' and kategori:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Laporan {kategori.upper()}"
        ws.append(['Tanggal', 'Nama Siswa', 'Jumlah (Rp)'])

        for cell in ws["1:1"]:
            cell.font = Font(bold=True)

        def write_rows(qs, field, tanggal_field, siswa_getter):
            for item in qs:
                siswa = siswa_getter(item)
                nama = getattr(siswa, 'nama', '-') if siswa else '-'
                jumlah = getattr(item, field, 0)
                tanggal = getattr(item, tanggal_field, tanggal_hari_ini)
                ws.append([tanggal, nama, float(jumlah)])

        if kategori == 'spp':
            write_rows(spp, 'jumlah_bayar', 'tanggal_bayar', lambda x: x.siswa)
        elif kategori == 'dsp':
            write_rows(dsp, 'jumlah', 'tanggal', lambda x: x.dsp.siswa)
        elif kategori == 'ppdb':
            write_rows(ppdb, 'jumlah', 'tanggal', lambda x: x.ppdb.siswa)
        elif kategori == 'tabungan':
            write_rows(tabungan, 'nominal', 'tanggal', lambda x: x.siswa)
        elif kategori == 'penarikan':
            write_rows(penarikan, 'jumlah_ditarik', 'tanggal', lambda x: x.siswa)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=laporan_{kategori}.xlsx'
        wb.save(response)
        return response

    context = {
        'tanggal': tanggal_hari_ini,
        'spp_total': spp.aggregate(Sum('jumlah_bayar'))['jumlah_bayar__sum'] or 0,
        'dsp_total': dsp.aggregate(Sum('jumlah'))['jumlah__sum'] or 0,
        'ppdb_total': ppdb.aggregate(Sum('jumlah'))['jumlah__sum'] or 0,
        'tabungan_total': tabungan.aggregate(Sum('nominal'))['nominal__sum'] or 0,
        'penarikan_total': penarikan.aggregate(Sum('jumlah_ditarik'))['jumlah_ditarik__sum'] or 0,
        'total_masuk': total_masuk,
        'total_keluar': total_keluar,
        'saldo': total_masuk - total_keluar,
        'transaksi': {
            'spp': spp,
            'dsp': dsp,
            'ppdb': ppdb,
            'tabungan': tabungan,
            'penarikan': penarikan,
        },
    }
    return render(request, 'keuangan/admin/laporan_keuangan.html', context)

def export_laporankeuangan_excel(request):
    hari_ini = datetime.date.today()

    spp = PembayaranSPP.objects.filter(tanggal__date=hari_ini, is_posted_to_laporan_keuangan=True)
    dsp = DSPCicilan.objects.filter(tanggal__date=hari_ini, is_posted_to_laporan_keuangan=True)
    ppdb = PPDBCicilan.objects.filter(tanggal__date=hari_ini, is_posted_to_laporan_keuangan=True)
    tabungan = TabunganSiswa.objects.filter(tanggal__date=hari_ini, is_posted_to_laporan_keuangan=True)
    penarikan = PenarikanTabungan.objects.filter(tanggal__date=hari_ini, is_posted_to_laporan_keuangan=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Keuangan"

    ws.append(["Tanggal", str(hari_ini)])
    ws.append([])
    ws.append(["Kategori", "Nama Siswa", "Jumlah (Rp)"])

    # Helper function
    def append_data(kategori, queryset, get_nama, get_jumlah):
        for obj in queryset:
            nama = get_nama(obj)
            jumlah = get_jumlah(obj)
            ws.append([kategori, nama, jumlah])

    append_data("SPP", spp, lambda x: x.siswa.nama, lambda x: x.jumlah_bayar)
    append_data("DSP", dsp, lambda x: x.dsp.siswa.nama, lambda x: x.jumlah)
    append_data("PPDB", ppdb, lambda x: x.ppdb.siswa.nama, lambda x: x.jumlah)
    append_data("Tabungan", tabungan, lambda x: x.siswa.nama, lambda x: x.nominal)
    append_data("Penarikan Tabungan", penarikan, lambda x: x.siswa.nama, lambda x: x.jumlah_ditarik)

    # Format kolom
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 2

    # Output response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = f"Laporan_Keuangan_{hari_ini}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response